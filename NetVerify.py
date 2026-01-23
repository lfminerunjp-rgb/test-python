import os, sys, json, re, subprocess, time, difflib, ctypes, binascii
from datetime import datetime
from collections import defaultdict
import math

# GUI Library
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QPushButton, QTableWidget, QTableWidgetItem, QTextEdit, 
                             QHeaderView, QComboBox, QLabel, QCheckBox, QGroupBox, 
                             QLineEdit, QFileDialog, QInputDialog, QTabWidget, QMessageBox, QScrollArea)
from PySide6.QtCore import Qt, QThread, Signal, Slot
from PySide6.QtGui import QTextCursor, QColor, QWheelEvent, QShortcut, QKeySequence, QDesktopServices 
from PySide6.QtCore import QUrl

import openpyxl
from netmiko import ConnectHandler

# Windows Registry (for TeraTerm detection)
try:
    import winreg
except ImportError:
    winreg = None

# --- Graph Library (for Mode 6 & 8) ---
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
try:
    import networkx as nx
    HAS_NETWORKX = True
except ImportError:
    HAS_NETWORKX = False

# --- 設定 ---
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

SNAPSHOT_DIR = os.path.join(BASE_DIR, "snapshots")
LOG_DIR = os.path.join(BASE_DIR, "logs")
REPORT_DIR = os.path.join(BASE_DIR, "reports")
PCAP_DIR = os.path.join(BASE_DIR, "pcaps")
SEARCH_FILE = os.path.join(BASE_DIR, "search.txt")

for d in [SNAPSHOT_DIR, LOG_DIR, REPORT_DIR, PCAP_DIR]:
    os.makedirs(d, exist_ok=True)

# --- モード6(自動診断)用: マルチベンダー対応設定 ---
DIAG_FAMILIES = {
    "cisco": [
        "cisco", "cisco_ios", "cisco_xe", "cisco_nxos", "cisco_xr", "cisco_asa", "cisco_wlc", 
        "arista_eos", "dell_force10", "dell_os10", "brocade_fastiron", "ruckus_fastiron", "icx_fastiron"
    ],
    "juniper": ["juniper", "juniper_junos", "juniper_screenos", "paloalto_panos", "vyos", "vyatta_vyos"],
    "huawei": [
        "huawei", "huawei_vrp", "huawei_smartax", "hp_comware", "h3c_comware", 
        "alcatel_sros", "nokia_sros", "alaxala"
    ],
    "hp_aruba": ["hp_procurve", "aruba_os", "aruba_osswitch", "aruba_aoscx", "aruba_procurve"],
    "fortinet": ["fortinet"],
    "yamaha": ["yamaha"],
    "allied": ["allied_telesis_awplus"],
    "nec": ["nec_ix"],
    "linux": ["linux", "linux_ssh", "ovs_linux", "redhat", "centos", "ubuntu", "debian"]
}

DIAG_COMMANDS = {
    "cisco": {
        "route": "show ip route {target}",
        "route_vrf": "show ip route vrf * {target}",
        "mpls": "show mpls forwarding-table {target}",
        "interface": "show interface {iface}",
        "config_if": "show run interface {iface}",
        "stp": "show spanning-tree interface {iface}",
        "channel": "show etherchannel summary",
        "arp": "show ip arp {next_hop}",
        "neighbor": "show cdp neighbors detail && show lldp neighbors detail",
        "acl": "show ip access-lists",
        "nat_conf": "show run | include ip nat",
        "nat_log": "show ip nat translations | include {target}",
        "ospf": "show ip ospf neighbor",
        "bgp": "show ip bgp summary",
        "eigrp": "show ip eigrp neighbors"
    },
    "juniper": {
        "route": "show route {target}",
        "route_vrf": "show route instance all {target}",
        "interface": "show interfaces {iface} detail",
        "config_if": "show configuration interfaces {iface}",
        "arp": "show arp hostname {next_hop}",
        "neighbor": "show lldp neighbors detail",
        "acl": "show firewall",
        "ospf": "show ospf neighbor",
        "bgp": "show bgp summary",
    },
    "huawei": {
        "route": "display ip routing-table {target}",
        "route_vrf": "display ip routing-table vpn-instance * {target}",
        "interface": "display interface {iface}",
        "config_if": "display current-configuration interface {iface}",
        "arp": "display arp | include {next_hop}",
        "neighbor": "display lldp neighbor verbose",
        "acl": "display acl all",
        "ospf": "display ospf peer",
        "bgp": "display bgp peer",
    },
    "hp_aruba": {
        "route": "show ip route {target}",
        "interface": "show interface {iface}",
        "config_if": "show running-config interface {iface}",
        "arp": "show arp | include {next_hop}",
        "neighbor": "show lldp info remote-device detail",
        "acl": "show access-list",
        "ospf": "show ip ospf neighbor",
        "bgp": "show ip bgp summary",
    },
    "fortinet": {
        "route": "get router info routing-table details {target}",
        "route_vrf": "get router info routing-table all {target}",
        "interface": "diagnose hardware deviceinfo nic {iface}",
        "config_if": "show system interface {iface}",
        "arp": "get system arp | grep {next_hop}",
        "neighbor": "get system lldp neighbor-summary",
        "acl": "show firewall policy",
        "ospf": "get router info ospf neighbor",
        "bgp": "get router info bgp summary",
    },
    "yamaha": {
        "route": "show ip route detail",
        "interface": "show status lan{iface}", 
        "config_if": "show config | grep lan{iface}",
        "arp": "show arp",
        "neighbor": "show lldp neighbor",
        "acl": "show ip filter",
    },
    "allied": {
        "route": "show ip route {target}",
        "interface": "show interface {iface}",
        "config_if": "show running-config interface {iface}",
        "arp": "show ip arp {next_hop}",
        "neighbor": "show lldp neighbors detail",
        "acl": "show ip access-list",
    },
    "nec": {
        "route": "show ip route {target}",
        "interface": "show interface {iface}",
        "config_if": "show running-config interface {iface}",
        "arp": "show ip arp",
        "neighbor": "show lldp neighbors detail",
        "acl": "show ip access-list",
        "ospf": "show ip ospf neighbor",
        "bgp": "show ip bgp summary",
    },
    "linux": {
        "route": "ip route get {target}",
        "interface": "ip addr show {iface}",
        "arp": "ip neigh show",
    }
}

# --- グラフ描画用キャンバス ---
class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=8, height=5, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi, facecolor='#1E1E1E')
        self.axes = self.fig.add_subplot(111)
        self.axes.set_facecolor('#1E1E1E')
        self.axes.tick_params(colors='white')
        for spine in self.axes.spines.values():
            spine.set_edgecolor('#555')
        self.axes.grid(True, color='#333', linestyle='--')
        
        self.current_dpi = dpi
        super().__init__(self.fig)

    def wheelEvent(self, event: QWheelEvent):
        if event.modifiers() & Qt.ControlModifier:
            if event.angleDelta().y() > 0:
                self.current_dpi += 5
            else:
                self.current_dpi = max(50, self.current_dpi - 5)
            
            self.fig.set_dpi(self.current_dpi)
            self.fig.set_size_inches(self.fig.get_size_inches()) 
            self.draw()
            event.accept()
        else:
            super().wheelEvent(event)

# --- カスタムTextEdit ---
class ZoomableTextEdit(QTextEdit):
    def wheelEvent(self, event: QWheelEvent):
        if event.modifiers() & Qt.ControlModifier:
            if event.angleDelta().y() > 0: self.zoomIn(1)
            else: self.zoomOut(1)
            event.accept()
        else: super().wheelEvent(event)

# --- 機能関数 ---
def sanitize_filename(name):
    return re.sub(r'[\\/:*?"<>|]', '_', str(name))

def find_teraterm(parent_widget=None):
    standard_paths = [r"C:\Program Files (x86)\teraterm\ttpmacro.exe", r"C:\Program Files\teraterm\ttpmacro.exe", r"C:\teraterm\ttpmacro.exe"]
    for p in standard_paths:
        if os.path.exists(p): return p
    if winreg:
        try:
            with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\ttermpro.exe") as key:
                path, _ = winreg.QueryValueEx(key, "")
                macro_path = os.path.join(os.path.dirname(path), "ttpmacro.exe")
                if os.path.exists(macro_path): return macro_path
        except: pass
    try:
        res = subprocess.check_output('where ttpmacro.exe', shell=True, stderr=subprocess.DEVNULL)
        path = res.decode('cp932').splitlines()[0].strip()
        if os.path.exists(path): return path
    except: pass
    return None

def clean_text_for_diff(text):
    if not isinstance(text, str): return []
    ignore_patterns = [r"^Current configuration\s*:\s*\d+\s*bytes", r"^Building configuration.*", r"^Last configuration change at.*", r"^NVRAM config last updated at.*", r"^!.*", r"^\s*$"]
    return [l for l in text.splitlines() if not any(re.search(p, l, re.IGNORECASE) for p in ignore_patterns)]

def generate_side_by_side_html(old_lines, new_lines, cmd_name=""):
    diff = list(difflib.ndiff(old_lines, new_lines))
    rows, p_left, p_right = [], [], []
    def flush():
        for i in range(max(len(p_left), len(p_right))):
            l_v, r_v = p_left[i] if i < len(p_left) else None, p_right[i] if i < len(p_right) else None
            if l_v is not None and r_v is None:
                rows.append(f'<tr><td style="color:#ff5555; background-color:#3a1a1a; border-left: 3px solid #ff5555;">{l_v}</td><td style="background-color:#222;">&nbsp;</td></tr>')
            elif r_v is not None and l_v is None:
                rows.append(f'<tr><td style="background-color:#222;">&nbsp;</td><td style="color:#5555ff; background-color:#1a1a3a; border-left: 3px solid #5555ff;">{r_v}</td></tr>')
            else:
                rows.append(f'<tr><td style="color:#ff5555; background-color:#3a1a1a;">{l_v}</td><td style="color:#5555ff; background-color:#1a1a3a;">{r_v}</td></tr>')
        p_left.clear(); p_right.clear()
    for line in diff:
        pre, con = line[:2], line[2:]
        if pre == '  ': flush(); rows.append(f'<tr><td>{con}</td><td>{con}</td></tr>')
        elif pre == '- ': p_left.append(con)
        elif pre == '+ ': p_right.append(con)
    flush()
    title_html = f'<div style="color:#FFFF00; font-weight:bold; margin-top:15px; text-align:left;">[差分あり] {cmd_name}</div>'
    table_html = f'<table border="0" width="100%" style="border-collapse:collapse; font-family:Consolas, monospace; color:#DDD; background:#1E1E1E; margin-left:0;"><tr style="background-color:#004d4d; color:#FFF;"><th>[ 前回 ]</th><th>[ 今回 ]</th></tr>{"".join(rows)}</table>'
    return title_html + table_html

# --- 処理スレッド (NetworkWorker) ---
class NetworkWorker(QThread):
    log_signal = Signal(str, str, str)
    html_signal = Signal(str, str)
    finished_signal = Signal(str, list, dict)
    request_teraterm_path = Signal() 

    def __init__(self, mode, host, show_output, scan_keywords, keywords_list, mesh_targets=None, compare_master=False, save_as_master=False, tt_path=None):
        super().__init__()
        self.mode, self.host = mode, host
        self.mesh_targets = mesh_targets or []
        self.compare_master, self.save_as_master = compare_master, save_as_master
        self.keywords_list, self.tt_path = keywords_list, tt_path
        self.report_data, self.mesh_results = [], {}
        self.current_process = None # プロセス制御用
        self._is_cancelled = False # キャンセル制御フラグ
        if "3:" in mode: self.show_output, self.scan_keywords = False, False
        else: self.show_output, self.scan_keywords = show_output, scan_keywords

    def stop(self):
        """スレッドを安全に停止させる"""
        self._is_cancelled = True
        self.kill_subprocess()

    def run(self):
        # 実行前にキャンセルチェック
        if self._is_cancelled: return

        today = datetime.now().strftime("%Y%m%d")
        h = self.host; name = h['name']
        self.log_signal.emit(name, f"\n{'='*25} {name} 開始 {'='*25}", "#FFFFFF")
        try:
            if "0:" in self.mode: self.do_ping(h, False)
            elif "0t:" in self.mode: self.do_ping(h, True)
            elif "1:" in self.mode: self.do_login(h)
            elif "5:" in self.mode: self.do_full_mesh_ping(h)
            else: self.do_netmiko(h, today)
        except Exception as e:
            self.log_signal.emit(name, f"[!] エラー: {str(e)}", "#FF5555")
        
        # 完了シグナル
        self.finished_signal.emit(name, self.report_data, self.mesh_results)

    def do_ping(self, h, trace):
        if self._is_cancelled: return
        ip, name = h.get('ip'), h['name']
        
        # Ping
        res = subprocess.run(['ping', '-n', '1', '-w', '1000', ip], stdout=subprocess.DEVNULL)
        status = "SUCCESS" if res.returncode == 0 else "FAIL"
        self.log_signal.emit(name, f"[{status}] Ping: {ip}", "#00FF00" if res.returncode == 0 else "#FF5555")
        
        # Trace
        if trace and not self._is_cancelled:
            self.log_signal.emit(name, "Tracerouteを実行中...", "#888888")
            cmd = ['tracert', '-d', '-h', '15', '-w', '200', ip] if os.name == 'nt' else ['traceroute', '-n', '-m', '15', ip]
            try:
                startupinfo = None
                if os.name == 'nt':
                    startupinfo = subprocess.STARTUPINFO()
                    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                
                # プロセス開始
                self.current_process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding='cp932' if os.name == 'nt' else 'utf-8', errors='replace', startupinfo=startupinfo)
                
                while not self._is_cancelled:
                    if self.current_process is None: break
                    try:
                        # 読み取り試行
                        line = self.current_process.stdout.readline()
                    except (ValueError, OSError):
                        # プロセスが外部から殺された場合のエラー回避
                        break
                    
                    if not line:
                        # データがなく、かつプロセスが終了していれば抜ける
                        if self.current_process.poll() is not None: break
                        continue
                        
                    self.log_signal.emit(name, line.rstrip(), "#FFFFFF")
            except Exception as e:
                self.log_signal.emit(name, f"Trace Error: {str(e)}", "#FF5555")
            finally:
                self.kill_subprocess()

    def kill_subprocess(self):
        if self.current_process:
            try:
                self.current_process.kill()
            except:
                pass
            self.current_process = None

    def do_login(self, h):
        if self._is_cancelled: return
        exe = self.tt_path if self.tt_path else find_teraterm()
        name = h['name']
        if not exe: 
            self.request_teraterm_path.emit()
            self.log_signal.emit(name, "TeraTermが見つかりません。ファイル選択画面で ttpmacro.exe を指定してください。", "#FF5555")
            return
        macro_path = os.path.join(BASE_DIR, f"temp_{sanitize_filename(name)}.ttl")
        with open(macro_path, "w", encoding='cp932') as f:
            if 'telnet' in str(h.get('protocol','')).lower(): 
                f.write(f"connect '{h['ip']}:23 /nossh /T=1'\nwait 'Username:' 'login:'\nsendln '{h.get('user') or ''}'\nwait 'Password:'\nsendln '{h.get('pw') or ''}'\n")
            else: 
                f.write(f"connect '{h['ip']}:22 /ssh /2 /auth=password /user={h.get('user') or ''} /passwd={h.get('pw') or ''}'\n")
            f.write(f"wait '>' '#'\nif result=1 then\nsendln 'enable'\nwaitregex '[Pp]assword|パスワード|暗号'\nsendln '{h.get('en_pw') or ''}'\nendif\n")
        subprocess.Popen([exe, macro_path])
        self.log_signal.emit(name, "TeraTermマクロ起動完了", "#00FF00")

    def do_netmiko(self, h, today):
        if self._is_cancelled: return
        name = h['name']; v, p = str(h.get('vendor') or 'cisco_ios').strip().lower(), str(h.get('protocol') or 'ssh').strip().lower()
        dev = {'device_type': v + ('_telnet' if p == 'telnet' else ''), 'host': h['ip'], 'username': h['user'], 'password': h['pw'], 'secret': h['en_pw'], 'global_delay_factor': 2}
        with ConnectHandler(**dev) as net:
            if ">" in net.find_prompt(): net.enable()
            self.check_save_status(net, name, v)
            outputs, log_body = {}, f"\n! --- Log: {datetime.now()} ---\n"
            for cmd in h.get('command_list', []):
                if self._is_cancelled: break # コマンドループもキャンセル可能に
                self.log_signal.emit(name, f"Command: {cmd}", "#AAAAAA")
                out = net.send_command(cmd, strip_prompt=True, strip_command=True)
                if self.scan_keywords and self.keywords_list:
                    for kw in self.keywords_list:
                        if kw and kw.lower() in out.lower():
                            for line in out.splitlines():
                                if kw.lower() in line.lower():
                                    self.html_signal.emit(name, f'<div style="color:#FFFF00; white-space:pre-wrap; font-family:Consolas; text-align:left;">    [HIT] \'{kw}\': {line.strip()}</div>')
                if self.show_output: 
                    safe_out = out.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    self.html_signal.emit(name, f'<div style="color:#FFFFFF; white-space:pre-wrap; font-family:Consolas; text-align:left;">{safe_out}</div>')
                outputs[cmd], log_body = out, log_body + f"{out}\n\n"
            if "解析" in self.mode or "比較" in self.mode: self.do_compare(name, outputs, h.get('command_list', []))
            if "2:" in self.mode or "4:" in self.mode:
                f_p = os.path.join(LOG_DIR, f"{sanitize_filename(name)}_{today}.log")
                with open(f_p, "a", encoding='utf-8') as f: f.write(log_body)
                self.log_signal.emit(name, f"[Log Saved] {os.path.basename(f_p)}", "#00AAFF")

    def check_save_status(self, net, name, vendor):
        self.log_signal.emit(name, "Running vs Startup 照合中...", "#888888")
        try:
            run_cmd, sta_cmd = "show running-config", "show startup-config"
            if "junos" in vendor: run_cmd, sta_cmd = "show configuration", "show configuration | display set"
            elif "fortigate" in vendor: run_cmd, sta_cmd = "show", "get system status"
            elif "yamaha" in vendor: run_cmd, sta_cmd = "show config", "show config"
            elif "arista" in vendor: run_cmd, sta_cmd = "show running-config", "show startup-config"
            elif "allied" in vendor: run_cmd, sta_cmd = "show running-config", "show startup-config"
            elif "nec" in vendor: run_cmd, sta_cmd = "show running-config", "show config"
            
            run = net.send_command(run_cmd); sta = net.send_command(sta_cmd)
            if clean_text_for_diff(run) != clean_text_for_diff(sta): self.log_signal.emit(name, "[!] 警告: 保存されていない設定があります", "#FF5555")
            else: self.log_signal.emit(name, "[OK] 設定保存済み", "#00FF00")
        except: pass

    def do_full_mesh_ping(self, h):
        if self._is_cancelled: return
        name, vendor = h['name'], str(h.get('vendor') or 'cisco_ios').strip().lower()
        dev = {'device_type': vendor + ('_telnet' if str(h.get('protocol')).lower() == 'telnet' else ''), 'host': h['ip'], 'username': h['user'], 'password': h['pw'], 'secret': h['en_pw']}
        ping_map = {
            "cisco": {"cmd": "ping {ip} repeat 2 timeout 1", "ok": "Success rate is 100"},
            "junos": {"cmd": "ping {ip} count 2 wait 1", "ok": "0% packet loss"},
            "fortigate": {"cmd": "execute ping {ip}", "ok": "0% packet loss"},
            "yamaha": {"cmd": "ping {ip} count 2", "ok": "Received from"},
            "aruba": {"cmd": "ping {ip} count 2", "ok": "0% packet loss"},
            "hp": {"cmd": "ping {ip} count 2", "ok": "is alive"},
            "allied": {"cmd": "ping {ip} count 2", "ok": "received"},
            "arista": {"cmd": "ping {ip} repeat 2", "ok": "100% packet loss"},
            "nec": {"cmd": "ping {ip} count 2", "ok": "100% packet loss"}
        }
        logic = next((v for k, v in ping_map.items() if k in vendor), ping_map["cisco"])
        with ConnectHandler(**dev) as net:
            if ">" in net.find_prompt(): net.enable()
            for t in self.mesh_targets:
                if self._is_cancelled: break
                if t['ip'] == h['ip']: self.mesh_results[t['name']] = "SELF"; continue
                self.log_signal.emit(name, f"Ping -> {t['name']}({t['ip']})", "#AAAAAA")
                res = net.send_command(logic["cmd"].format(ip=t['ip']))
                
                is_ok = False
                if "packet loss" in logic["ok"]:
                    if "0% packet loss" in res: is_ok = True
                elif logic["ok"] in res:
                    is_ok = True
                
                self.mesh_results[t['name']] = "OK" if is_ok else "NG"
                self.log_signal.emit(name, f"  result: {'OK' if is_ok else 'NG'}", "#00FF00" if is_ok else "#FF5555")

    def do_compare(self, name, current, cmds):
        h_file = sanitize_filename(name)
        target = os.path.join(SNAPSHOT_DIR, f"snapshot_{h_file}_master.json") if self.compare_master else os.path.join(SNAPSHOT_DIR, f"snapshot_{h_file}.json")
        if os.path.exists(target):
            self.log_signal.emit(name, f"[Compare] 使用ファイル: {os.path.basename(target)}", "#00AAFF")
            try:
                with open(target, "r", encoding='utf-8') as f: old = json.load(f)
            except Exception as e:
                self.log_signal.emit(name, f"[!] 比較元ファイル読み込み失敗: {e}", "#FF5555"); return
            self.report_data.append(f'<h2 style="color:#00FFFF; border-bottom:2px solid #00FFFF; text-align:left;">Device: {name} (比較対象: {os.path.basename(target)})</h2>')
            diff_count = 0
            for cmd in cmds:
                if cmd not in old:
                    msg = f'<div style="color:#FFFF00;">[新規取得] {cmd} が比較元に存在しません。</div>'
                    self.html_signal.emit(name, msg); self.report_data.append(msg); diff_count += 1; continue
                old_lines = clean_text_for_diff(old[cmd]); new_lines = clean_text_for_diff(current[cmd])
                if old_lines != new_lines:
                    h_res = generate_side_by_side_html(old_lines, new_lines, cmd)
                    self.html_signal.emit(name, h_res); self.report_data.append(h_res); diff_count += 1
            if diff_count == 0:
                no_diff_msg = f'<div style="color:#00FF00; margin-top:10px; font-family:Consolas;">    [Result] 差分なし (Config is synced)</div>'
                self.html_signal.emit(name, no_diff_msg); self.report_data.append(no_diff_msg)
                self.log_signal.emit(name, "    [Result] 差分なし (前回のスナップショットと同じです)", "#00FF00")
        else: self.log_signal.emit(name, f"[Compare] 比較対象なし (新規スナップショットとして扱います)", "#AAAAAA")
        if self.save_as_master:
            snap_path = os.path.join(SNAPSHOT_DIR, f"snapshot_{h_file}_master.json")
            with open(snap_path, "w", encoding='utf-8') as f: json.dump(current, f, indent=4, ensure_ascii=False)
            self.log_signal.emit(name, "[OK] Masterファイルとして保存しました。", "#00FF00")
        else:
            snap_path = os.path.join(SNAPSHOT_DIR, f"snapshot_{h_file}.json")
            if os.path.exists(snap_path):
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                try: os.rename(snap_path, os.path.join(SNAPSHOT_DIR, f"snapshot_{h_file}_{ts}.json"))
                except: pass
            with open(snap_path, "w", encoding='utf-8') as f: json.dump(current, f, indent=4, ensure_ascii=False)
            self.log_signal.emit(name, "[OK] スナップショットを作成（更新）しました。", "#00FF00")

# --- モード6用: CCIEレベル自動診断ワーカー (DiagnosticWorker) ---
class DiagnosticWorker(QThread):
    log_signal = Signal(str, str, str); finished_signal = Signal(str, list, dict)
    path_data_signal = Signal(list) # 経路可視化用データ

    def __init__(self, start, target, hosts):
        super().__init__(); self.cur, self.tgt, self.hosts = start, target, hosts; self.rep, self.visited = [], set()
        self.path_trace = [] # 経路可視化用

    def run(self):
        self.log_signal.emit("DIAG", f"=== 自動診断開始: {self.cur['name']} -> {self.tgt} ===", "#00FFFF")
        hop, found = 0, False
        self.path_trace.append({"node": "START_PC", "next": self.cur['name'], "iface": "access", "status": "OK", "reason": ""})
        
        while hop < 15 and not found:
            hop += 1; h = self.cur; n, ip = h['name'], h['ip']
            self.log_signal.emit(n, f"--- Hop {hop}: {n} ({ip}) ---", "#00FF00")
            
            try:
                v = str(h.get('vendor') or 'cisco_ios').lower()
                v_fam = "cisco"
                for fam, drivers in DIAG_FAMILIES.items():
                    if any(d in v for d in drivers): v_fam = fam; break
                cmds = DIAG_COMMANDS.get(v_fam, DIAG_COMMANDS["cisco"])
                
                dev = {'device_type': v, 'host': ip, 'username': h['user'], 'password': h['pw'], 'secret': h['en_pw'], 'global_delay_factor': 2}
                with ConnectHandler(**dev) as net:
                    if ">" in net.find_prompt(): net.enable()

                    # L3 Routing
                    rout = net.send_command(cmds["route"].format(target=self.tgt))
                    if "not in table" in rout and "route_vrf" in cmds:
                        try:
                            vrout = net.send_command(cmds["route_vrf"].format(target=self.tgt))
                            if "not in table" not in vrout and vrout.strip(): rout = vrout
                        except: pass
                    self.rep.append(f"<h3>[{n}] Route ({v_fam})</h3><pre>{rout}</pre>")
                    
                    nh_ip = self.get_nh(rout, v_fam); iface = self.get_iface(rout)
                    
                    path_node = {"node": n, "next": nh_ip, "iface": iface, "status": "OK", "reason": ""}

                    if not nh_ip and not iface:
                        self.log_signal.emit(n, "[!] Route Missing (Drop)", "#FF5555")
                        self.rep.append(f"<div style='color:red'><b>[CAUSE] Routing Missing at {n}</b></div>")
                        path_node["status"] = "FAIL"
                        path_node["reason"] = "No Route"
                        self.path_trace.append(path_node); found = True; break
                    
                    if any(x in rout.lower() for x in ["connected", "direct", "via 0.0.0.0", "attached", "is directly"]):
                        self.log_signal.emit(n, f"[Info] Directly Connected ({iface})", "#FFFF00"); nh_ip = self.tgt

                    if iface:
                        iout = net.send_command(cmds["interface"].format(iface=iface))
                        self.rep.append(f"<h3>[{n}] Interface {iface}</h3><pre>{iout}</pre>")
                        
                        if "err-disabled" in iout.lower():
                            self.log_signal.emit(n, f"[Critical] Port {iface} is ERR-DISABLED!", "#FF0000")
                            path_node["status"] = "FAIL"
                            path_node["reason"] = "Err-Disabled"
                            self.path_trace.append(path_node); found = True; break
                        if "down" in iout.lower():
                            self.log_signal.emit(n, f"[!] Port {iface} is DOWN", "#FF5555")
                            path_node["status"] = "FAIL"
                            path_node["reason"] = "Link Down"
                            self.path_trace.append(path_node); found = True; break
                        
                        self.check_if_quality(n, iout)

                    self.path_trace.append(path_node)

                    if nh_ip == self.tgt:
                        self.log_signal.emit(n, "[End] 到達しました。端末FW等を確認してください。", "#00FF00"); found = True; break

                    self.log_signal.emit(n, f"[Next] -> {nh_ip} (via {iface})", "#FFFFFF")
                    next_h = self.find_host(nh_ip)
                    if not next_h:
                        if "arp" in cmds:
                            try:
                                arp = net.send_command(cmds["arp"].format(next_hop=nh_ip))
                                if any(m in arp for m in ["0000.0c07.ac", "0000.5e00.01", "0000.0c9f.f"]): self.log_signal.emit(n, f"[Info] {nh_ip} is Virtual IP (VIP).", "#00FFFF")
                            except: pass
                        res = net.send_command(f"ping {nh_ip}" if "cisco" in v_fam else f"ping {nh_ip} count 2")
                        if not any(s in res for s in ["!!!!", "100", "0% packet loss", "received", "alive"]):
                            self.log_signal.emit(n, f"[!] Ping NG to {nh_ip}. Link broken?", "#FF5555")
                            self.rep.append(f"<div style='color:red'><b>[CAUSE] Unreachable Next Hop {nh_ip}</b></div>")
                            self.path_trace[-1]["status"] = "FAIL"
                            self.path_trace[-1]["reason"] = "Ping NG"
                        found = True; break
                    
                    if next_h['ip'] in self.visited: 
                        self.log_signal.emit(n, "[!] Routing Loop Detected!", "#FF0000"); 
                        self.path_trace[-1]["status"] = "LOOP"
                        self.path_trace[-1]["reason"] = "Loop"
                        break
                    self.visited.add(next_h['ip']); self.cur = next_h
            except Exception as e:
                self.log_signal.emit(n, f"[Err] {e}", "#FF5555"); found = True; break
        
        # Ensure completion signal is sent even after break
        self.finished_signal.emit("DIAG", self.rep, {})
        self.path_data_signal.emit(self.path_trace)

    def check_if_quality(self, n, out):
        if re.search(r"drops?[:\s]+(\d+)", out, re.I):
             d = re.findall(r"drops?[:\s]+(\d+)", out, re.I)
             if any(int(x)>0 for x in d): self.log_signal.emit(n, "[Warn] Interface Drops detected!", "#FFA500")
        if "CRC" in out and re.search(r"(\d+)\s+CRC", out):
             if int(re.search(r"(\d+)\s+CRC", out).group(1)) > 0: self.log_signal.emit(n, "[Alert] CRC Errors detected!", "#FF0000")
        if "Half-duplex" in out: self.log_signal.emit(n, "[Alert] Half-Duplex detected!", "#FF0000")

    def get_nh(self, txt, fam):
        if fam in ["cisco", "aruba_procurve", "hp_aruba", "arista", "allied", "nec"]: m=re.search(r"via\s+(\d{1,3}(?:\.\d{1,3}){3})", txt); return m.group(1) if m else None
        if fam == "juniper": m=re.search(r"to\s+(\d{1,3}(?:\.\d{1,3}){3})", txt); return m.group(1) if m else None
        if fam == "huawei": m=re.search(r"RelayNextHop\s*:\s*(\d{1,3}(?:\.\d{1,3}){3})", txt); return m.group(1) if m else None
        m=re.search(r"(\d{1,3}(?:\.\d{1,3}){3})", txt); return m.group(1) if m else None

    def get_iface(self, txt):
        for p in [r"(GigabitEthernet[\d/]+)", r"(TenGigabitEthernet[\d/]+)", r"(FastEthernet[\d/]+)", r"(Eth[\d/]+)", 
                  r"(ge-[\d/\.]+)", r"(xe-[\d/\.]+)", r"(Vlan\d+)", r"(Port-channel\d+)", r"(Eth-Trunk\d+)", r"(Tunnel\d+)",
                  r"(ethernet[\d/]+)"]:
            m=re.search(p, txt, re.I); 
            if m: return m.group(1)
        return None
    
    def find_host(self, ip):
        for h in self.hosts:
            if h['ip'] == ip: return h
        return None

# --- モード7用: 帯域モニターワーカー (TrafficGhostWorker) ---
class TrafficGhostWorker(QThread):
    log_signal = Signal(str, str, str)
    update_signal = Signal(str, float, float)
    finished_signal = Signal(str, list, dict)

    def __init__(self, host, interface):
        super().__init__()
        self.host = host
        self.interface = interface
        self.is_running = True

    def run(self):
        h = self.host; name = h['name']
        v = str(h.get('vendor') or 'cisco_ios').strip().lower()
        dev = {'device_type': v, 'host': h['ip'], 'username': h['user'], 'password': h['pw'], 'secret': h['en_pw']}
        
        # Command Selection based on vendor
        cmd = f"show interface {self.interface}"
        if "juniper" in v: cmd = f"show interfaces {self.interface} detail"
        elif "huawei" in v: cmd = f"display interface {self.interface}"
        elif "fortinet" in v: cmd = f"diagnose hardware deviceinfo nic {self.interface}"
        elif "hp" in v or "aruba" in v: cmd = f"show interface {self.interface}"
        elif "allied" in v: cmd = f"show interface {self.interface}"
        elif "nec" in v: cmd = f"show interface {self.interface}"
        elif "arista" in v: cmd = f"show interface {self.interface}"
        
        last_in, last_out, last_time = None, None, None
        try:
            with ConnectHandler(**dev) as net:
                if ">" in net.find_prompt(): net.enable()
                self.log_signal.emit(name, f"--- [Ghost Mode] 監視開始: {self.interface} ---", "#00FFFF")
                while self.is_running:
                    now_time = datetime.now()
                    output = net.send_command(cmd)
                    mi=re.search(r"input,?\s+(\d+)\s+bytes|Input bytes\s*:\s*(\d+)|Rx\s+bytes:(\d+)", output, re.I|re.S)
                    mo=re.search(r"output,?\s+(\d+)\s+bytes|Output bytes\s*:\s*(\d+)|Tx\s+bytes:(\d+)", output, re.I|re.S)
                    if mi and mo:
                        curr_in = int(next(g for g in mi.groups() if g))
                        curr_out = int(next(g for g in mo.groups() if g))
                        if last_in is not None:
                            diff_time = (now_time - last_time).total_seconds()
                            if diff_time > 0:
                                mbps_in = ((curr_in - last_in) * 8 / diff_time) / 1_000_000
                                mbps_out = ((curr_out - last_out) * 8 / diff_time) / 1_000_000
                                self.update_signal.emit(now_time.strftime("%H:%M:%S"), max(0, mbps_in), max(0, mbps_out))
                        last_in, last_out, last_time = curr_in, curr_out, now_time
                    else: self.log_signal.emit(name, "[!] データ抽出失敗 (Regex Unmatched)", "#FF5555")
                    time.sleep(3)
        except Exception as e: self.log_signal.emit(name, f"[!] 接続エラー: {str(e)}", "#FF5555")
        self.finished_signal.emit(name, [], {})

    def stop(self): self.is_running = False

# --- モード8用: ネットワーククローラー (CrawlerWorker) ---
class CrawlerWorker(QThread):
    log_signal = Signal(str, str, str); finished_signal = Signal(str, list, dict)
    html_ready_signal = Signal(str) # HTML file path

    def __init__(self, start_host, hosts_data):
        super().__init__()
        self.start_host = start_host
        self.hosts_data = hosts_data
        self.visited = set()
        self.G = nx.Graph() if HAS_NETWORKX else None

    def run(self):
        if not HAS_NETWORKX:
            self.log_signal.emit("Crawler", "networkx がインストールされていません。", "#FF0000")
            self.finished_signal.emit("Crawler", [], {})
            return

        self.log_signal.emit("Crawler", f"Crawler Start from: {self.start_host['name']}", "#00FFFF")
        queue = [self.start_host]
        self.visited.add(self.start_host['ip'])
        
        # Data structures
        mac_db = {}
        arp_db = {} 
        host_mac_map = {} 
        
        host_ip_map = {h['ip']: h['name'] for h in self.hosts_data}

        # 1. データ収集フェーズ
        total = len(self.hosts_data)
        for idx, h in enumerate(self.hosts_data):
            self.log_signal.emit("Crawler", f"Scanning {h['name']} ({idx+1}/{total})...", "#AAAAAA")
            try:
                v = str(h.get('vendor') or 'cisco_ios').lower()
                dev = {'device_type': v, 'host': h['ip'], 'username': h['user'], 'password': h['pw'], 'secret': h['en_pw']}
                with ConnectHandler(**dev) as net:
                    if ">" in net.find_prompt(): net.enable()
                    
                    # Get ARP
                    arp_cmd = "show ip arp" 
                    if "junos" in v: arp_cmd = "show arp"
                    elif "huawei" in v: arp_cmd = "display arp"
                    elif "hp" in v or "aruba" in v: arp_cmd = "show arp"
                    elif "yamaha" in v: arp_cmd = "show arp"
                    elif "fortinet" in v: arp_cmd = "get system arp"

                    arp_out = net.send_command(arp_cmd)
                    for line in arp_out.splitlines():
                        m = re.search(r"(\d+\.\d+\.\d+\.\d+)\s+.*([0-9a-fA-F]{4}\.[0-9a-fA-F]{4}\.[0-9a-fA-F]{4})", line)
                        # MAC format variations (colon, dash) can be handled here if needed
                        if m: arp_db[m.group(1)] = m.group(2)
                    
                    # Get MAC Table
                    mac_cmd = "show mac address-table" 
                    if "junos" in v: mac_cmd = "show ethernet-switching table"
                    elif "huawei" in v: mac_cmd = "display mac-address"
                    elif "yamaha" in v: mac_cmd = "show switch mac address-table"
                    elif "hp" in v: mac_cmd = "show mac-address"
                    elif "allied" in v: mac_cmd = "show mac address-table"

                    mac_out = net.send_command(mac_cmd)
                    
                    # Identify Self MAC
                    if h['ip'] in arp_db:
                        host_mac_map[arp_db[h['ip']]] = h['name']
                    
                    h_macs = defaultdict(set)
                    for line in mac_out.splitlines():
                        parts = line.split()
                        if len(parts) >= 4:
                            # Basic heuristic for MAC and Port. Might need vendor-specific regex for robustness
                            mac_cand = parts[1]
                            port_cand = parts[-1]
                            # Attempt to find MAC in line
                            m_obj = re.search(r"([0-9a-fA-F]{4}[\.:-][0-9a-fA-F]{4}[\.:-][0-9a-fA-F]{4})", line)
                            if m_obj:
                                mac_cand = m_obj.group(1)
                                # Last word often port, but not always.
                                h_macs[parts[-1]].add(mac_cand)
                    mac_db[h['name']] = h_macs

            except Exception as e:
                self.log_signal.emit(h['name'], f"Scan Failed: {e}", "#FF5555")

        # 2. グラフ構築フェーズ (保護)
        try:
            self.G.add_nodes_from([h['name'] for h in self.hosts_data])
            self.log_signal.emit("Crawler", "Calculating Topology...", "#00AAFF")
            
            raw_links = []
            
            for h_a in self.hosts_data:
                name_a = h_a['name']
                if name_a not in mac_db: continue
                
                for port, macs in mac_db[name_a].items():
                    for mac in macs:
                        if mac in host_mac_map:
                            name_b = host_mac_map[mac]
                            if name_a != name_b:
                                raw_links.append((name_a, name_b, port))
            
            # Consolidate bi-directional links
            link_map = defaultdict(dict)
            for u, v, p in raw_links:
                key = tuple(sorted((u, v)))
                link_map[key][u] = p
                
            for (u, v), ports in link_map.items():
                port_u = ports.get(u, "?")
                port_v = ports.get(v, "?")
                label = f"{port_u} <--> {port_v}"
                self.G.add_edge(u, v, label=label)
                self.log_signal.emit("Crawler", f"Link: {u}[{port_u}] -- {v}[{port_v}]", "#00FF00")

            # HTML生成
            html_path = self.generate_html(self.G)
            self.html_ready_signal.emit(html_path)

        except Exception as e:
            self.log_signal.emit("Crawler", f"Topology Calculation Error: {e}", "#FF0000")
        
        finally:
            self.finished_signal.emit("Crawler", [], {})

    def generate_html(self, G):
        nodes = []
        edges = []
        for n in G.nodes():
            nodes.append({'id': n, 'label': n})
        for u, v, data in G.edges(data=True):
            edges.append({'from': u, 'to': v, 'label': data.get('label', ''), 'arrows': 'to;from'})
        
        html_content = """
<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>Network Topology</title>
  <script type="text/javascript" src="https://unpkg.com/vis-network/standalone/umd/vis-network.min.js"></script>
  <style type="text/css">
    #mynetwork { width: 100%%; height: 90vh; border: 1px solid lightgray; background-color: #222; }
    body { margin: 0; background-color: #111; color: #eee; font-family: sans-serif; }
    h2 { padding: 10px; margin: 0; display:inline-block; }
    .controls { padding: 10px; border-bottom: 1px solid #444; }
    button { background: #444; color: white; border: 1px solid #666; padding: 5px 10px; cursor: pointer; }
    button:hover { background: #555; }
  </style>
</head>
<body>
  <div class="controls">
    <h2>Network Topology</h2>
    <button onclick="togglePhysics()">Toggle Physics (Freeze/Unfreeze)</button>
  </div>
  <div id="mynetwork"></div>
  <script type="text/javascript">
    var nodes = new vis.DataSet(%s);
    var edges = new vis.DataSet(%s);
    var container = document.getElementById('mynetwork');
    var data = { nodes: nodes, edges: edges };
    var options = {
      nodes: { 
        shape: 'dot', size: 25, 
        font: { color: '#ffffff', size: 16 },
        borderWidth: 2, color: { background: '#008080', border: '#ffffff' }
      },
      edges: { 
        width: 2, 
        color: { color: '#aaaaaa', highlight: '#00FFFF' }, 
        font: { color: '#ffffff', size: 12, strokeWidth: 2, strokeColor: '#000000', align: 'top' },
        smooth: { type: 'continuous' }
      },
      physics: { 
        enabled: true,
        stabilization: { iterations: 1000 },
        barnesHut: { 
            gravitationalConstant: -3000, 
            centralGravity: 0.0,  
            springLength: 200, 
            springConstant: 0.01, 
            damping: 0.5          
        }
      },
      interaction: { hover: true, navigationButtons: true, keyboard: true }
    };
    var network = new vis.Network(container, data, options);
    network.on("dragStart", function (params) {
        if (params.nodes.length > 0) {
            var nodeId = params.nodes[0];
            nodes.update({id: nodeId, fixed: false});
        }
    });
    network.on("dragEnd", function (params) {
        if (params.nodes.length > 0) {
            var nodeId = params.nodes[0];
            nodes.update({id: nodeId, fixed: {x: false, y: false}});
        }
    });
    function togglePhysics() {
        options.physics.enabled = !options.physics.enabled;
        network.setOptions(options);
    }
  </script>
</body>
</html>
""" % (json.dumps(nodes), json.dumps(edges))
        
        path = os.path.join(REPORT_DIR, f"topology_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
        with open(path, "w", encoding='utf-8') as f:
            f.write(html_content)
        return path

# --- モード9用: 仮想ワイヤータップ (WiretapWorker) ---
class WiretapWorker(QThread):
    log_signal = Signal(str, str, str); finished_signal = Signal(str, list, dict)

    def __init__(self, host, interface, pcap_filter, duration):
        super().__init__()
        self.host, self.iface, self.filter, self.duration = host, interface, pcap_filter, duration

    def run(self):
        name = self.host['name']
        v = str(self.host.get('vendor') or 'cisco_ios').lower()
        self.log_signal.emit(name, f"--- Virtual Wiretap Start ({self.duration}s) ---", "#FF00FF")
        
        pcap_data = b""
        try:
            dev = {'device_type': v, 'host': self.host['ip'], 'username': self.host['user'], 'password': self.host['pw'], 'secret': self.host['en_pw']}
            with ConnectHandler(**dev) as net:
                if ">" in net.find_prompt(): net.enable()
                if "cisco" in v or "arista" in v or "allied" in v:
                    try:
                        net.send_command("no monitor capture point ip cef CAPPOINT", expect_string=r"#")
                        net.send_command("no monitor capture buffer CAPBUF", expect_string=r"#")
                        self.log_signal.emit(name, "Configuring EPC...", "#888")
                        net.send_command(f"monitor capture buffer CAPBUF size 2048 max-size 1518 linear")
                        filter_cmd = f"monitor capture point ip cef CAPPOINT {self.iface} both"
                        net.send_command(filter_cmd)
                        net.send_command("monitor capture point associate CAPPOINT CAPBUF")
                        net.send_command("monitor capture point start CAPPOINT")
                        self.log_signal.emit(name, f"Capturing for {self.duration} sec...", "#00FFFF")
                        time.sleep(self.duration)
                        net.send_command("monitor capture point stop CAPPOINT")
                        self.log_signal.emit(name, "Downloading Buffer...", "#00AAFF")
                        out = net.send_command("show monitor capture buffer CAPBUF dump")
                        pcap_data = self.parse_cisco_hex_dump(out)
                        net.send_command("no monitor capture point ip cef CAPPOINT")
                        net.send_command("no monitor capture buffer CAPBUF")
                    except Exception:
                        self.log_signal.emit(name, "[!] Capture command not supported on this device/version", "#FF5555")

                elif "linux" in v or "aruba_aoscx" in v or "vyos" in v:
                    self.log_signal.emit(name, "Running tcpdump...", "#00FFFF")
                    cmd_hex = f"timeout {self.duration} tcpdump -i {self.iface} -s 0 -x {self.filter}"
                    if "aruba" in v:
                        self.log_signal.emit(name, "[!] Aruba CX capture via CLI text dump is experimental", "#FFA500")
                        cmd_hex = f"diag utilities tcpdump -i {self.iface} -w -" 
                    
                    out = net.send_command(cmd_hex, read_timeout=self.duration + 10)
                    pcap_data = self.parse_tcpdump_hex(out)
                else:
                    self.log_signal.emit(name, f"[!] Wiretap not implemented for vendor: {v}", "#FF5555")

            if pcap_data:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                fname = os.path.join(PCAP_DIR, f"wiretap_{name}_{ts}.pcap")
                with open(fname, "wb") as f: f.write(pcap_data)
                self.log_signal.emit(name, f"[Success] Saved to {fname}", "#00FF00")
            elif "implemented" not in str(pcap_data):
                self.log_signal.emit(name, "[!] No packets captured or parsing failed", "#FFA500")

        except Exception as e:
            self.log_signal.emit(name, f"Wiretap Error: {str(e)}", "#FF5555")
        
        self.finished_signal.emit(name, [], {})

    def parse_cisco_hex_dump(self, text):
        data = bytearray()
        for line in text.splitlines():
            parts = line.split()
            if not parts: continue
            for p in parts:
                if re.match(r"^[0-9A-Fa-f]{4}$", p):
                    try: data.extend(binascii.unhexlify(p))
                    except: pass
        return data

    def parse_tcpdump_hex(self, text):
        data = bytearray()
        for line in text.splitlines():
            parts = line.split()
            for p in parts:
                if re.match(r"^[0-9A-Fa-f]{4}$", p):
                     try: data.extend(binascii.unhexlify(p))
                     except: pass
        return data


# --- GUI ---
class NetVerifyGUI(QMainWindow):
    def __init__(self):
        super().__init__(); self.setWindowTitle("NetVerify Pro - Professional Final Edition"); self.resize(1550, 950)
        self.hosts_data, self.active_workers, self.current_report_html, self.host_consoles, self.full_mesh_matrix = [], [], [], {}, {}
        self.ghost_x, self.ghost_in, self.ghost_out = [], [], []
        self.canvas = None
        self.teraterm_path = None 
        
        self.setup_ui(); self.load_excel(); self.setup_shortcuts()

    def setup_ui(self):
        cw = QWidget(); self.setCentralWidget(cw); main_layout = QHBoxLayout(cw); left_panel = QVBoxLayout()
        g_hosts = QGroupBox("機器選択"); ghl = QVBoxLayout(); g_hosts.setLayout(ghl)
        self.btn_all = QPushButton("全選択/解除"); self.btn_all.setFixedHeight(35); self.btn_all.clicked.connect(self.toggle_all)
        ghl.addWidget(self.btn_all); self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["選択", "ホスト名", "IPアドレス"]); self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        ghl.addWidget(self.table); left_panel.addWidget(g_hosts); main_layout.addLayout(left_panel, 1)

        right_panel = QVBoxLayout(); g_cfg = QGroupBox("実行コントロール"); gcl = QHBoxLayout(); g_cfg.setLayout(gcl)
        self.combo = QComboBox(); self.combo.setMinimumWidth(180); self.combo.setFixedHeight(35)
        
        # === モード ===
        self.combo.addItems([
            "0: Ping", "0t: Trace", "1: Login", "2: ログ取得", "3: 解析・比較", "4: ログ+比較", 
            "5: フルメッシュPing", "6: 自動診断 (Auto-Tshoot)", "7: 帯域モニター",
            "8: トポロジー自動描画 (Crawler)", "9: 仮想ワイヤータップ"
        ])
        
        self.combo.currentIndexChanged.connect(self.on_mode_changed)
        self.search_input = QLineEdit(); self.search_input.setPlaceholderText("ログから検索 (F3で次へ)..."); self.search_input.setFixedHeight(35); self.search_input.setStyleSheet("background:#333; color:white; border: 1px solid #555; padding-left:10px;")
        self.search_input.textChanged.connect(self.search_all_highlight); self.search_input.returnPressed.connect(self.find_next_match)
        
        btn_style = "QPushButton {{ font-weight: bold; font-size: 10pt; border-radius: 4px; color: white; background-color: {0}; }}"
        self.btn_run = QPushButton(" 実行開始 "); self.btn_run.setFixedSize(110, 40); self.btn_run.setStyleSheet(btn_style.format("#1a4d1a")); self.btn_run.clicked.connect(self.run_process)
        self.btn_cancel = QPushButton("キャンセル"); self.btn_cancel.setFixedSize(110, 40); self.btn_cancel.setEnabled(False); self.btn_cancel.setStyleSheet(btn_style.format("#4d1a1a")); self.btn_cancel.clicked.connect(self.stop_workers)
        self.btn_report = QPushButton("レポート保存"); self.btn_report.setFixedSize(110, 40); self.btn_report.setEnabled(False); self.btn_report.setStyleSheet(btn_style.format("#004d4d")); self.btn_report.clicked.connect(self.save_report)
        self.btn_clear = QPushButton("ログクリア"); self.btn_clear.setFixedSize(110, 40); self.btn_clear.setStyleSheet(btn_style.format("#333") + "border: 1px solid #666;"); self.btn_clear.clicked.connect(self.reset_all_logs_and_tabs)
        
        gcl.addWidget(QLabel("モード:")); gcl.addWidget(self.combo); gcl.addSpacing(10); gcl.addWidget(self.search_input); gcl.addSpacing(10); gcl.addWidget(self.btn_run); gcl.addWidget(self.btn_cancel); gcl.addWidget(self.btn_report); gcl.addWidget(self.btn_clear); right_panel.addWidget(g_cfg)

        option_layout = QHBoxLayout(); self.chk_show_log = QCheckBox("取得内容を画面に表示する"); self.chk_show_log.setVisible(False); self.chk_show_log.setChecked(True)
        self.chk_keyword_scan = QCheckBox("search.txt のキーワードを検知する"); self.chk_keyword_scan.setVisible(False); self.chk_keyword_scan.setChecked(True)
        self.chk_compare_master = QCheckBox("Masterと比較する"); self.chk_compare_master.setVisible(False); self.chk_compare_master.setStyleSheet("color: white; font-weight: bold;")
        self.chk_save_master = QCheckBox("Masterとして保存（更新）する"); self.chk_save_master.setVisible(False); self.chk_save_master.setStyleSheet("color: white; font-weight: bold;")
        option_layout.addWidget(self.chk_show_log); option_layout.addWidget(self.chk_keyword_scan); option_layout.addWidget(self.chk_compare_master); option_layout.addWidget(self.chk_save_master); option_layout.addStretch(); right_panel.addLayout(option_layout)
        
        self.tabs = QTabWidget(); self.tabs.setTabsClosable(True); self.tabs.tabCloseRequested.connect(lambda i: self.tabs.removeTab(i) if i != 0 else None)
        self.global_console = ZoomableTextEdit(); self.global_console.setReadOnly(True); self.global_console.setStyleSheet("background-color:#1E1E1E; color:#FFFFFF; font-family:Consolas, monospace; border:2px solid #333;")
        self.tabs.addTab(self.global_console, "全体ログ"); right_panel.addWidget(self.tabs); main_layout.addLayout(right_panel, 4)

    def setup_shortcuts(self):
        QShortcut(QKeySequence("F3"), self).activated.connect(self.find_next_match); QShortcut(QKeySequence("Ctrl+F"), self).activated.connect(self.search_input.setFocus)

    def search_all_highlight(self):
        text = self.search_input.text(); cur = self.tabs.currentWidget()
        if not isinstance(cur, ZoomableTextEdit): return
        if not text: cur.setExtraSelections([]); return
        extra = []; doc = cur.document(); cursor = QTextCursor(doc)
        while True:
            cursor = doc.find(text, cursor)
            if cursor.isNull(): break
            sel = QTextEdit.ExtraSelection(); sel.format.setBackground(QColor(144, 238, 144, 150)); sel.format.setForeground(Qt.black); sel.cursor = cursor; extra.append(sel)
        cur.setExtraSelections(extra)

    def find_next_match(self):
        text = self.search_input.text(); cur = self.tabs.currentWidget()
        if not isinstance(cur, ZoomableTextEdit) or not text: return
        if not cur.find(text): cur.moveCursor(QTextCursor.Start); cur.find(text)

    def on_mode_changed(self, index):
        mode = self.combo.currentText()
        show_content_opts = "2:" in mode or "4:" in mode
        self.chk_show_log.setVisible(show_content_opts)
        self.chk_keyword_scan.setVisible(show_content_opts)
        is_master_mode = "3:" in mode or "4:" in mode
        self.chk_compare_master.setVisible(is_master_mode)
        self.chk_save_master.setVisible(is_master_mode)

    def load_excel(self):
        xlsx_path = os.path.join(BASE_DIR, "inventory.xlsx")
        if not os.path.exists(xlsx_path): return
        wb = openpyxl.load_workbook(xlsx_path, data_only=True); ws = wb.active; headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if not row or row[0] is None: continue
            h = {headers[j]: row[j] for j, n in enumerate(headers) if n and j < len(row)}
            h['en_pw'] = row[6] if len(row)>6 else ""; h['command_list'] = [c.strip() for c in str(row[7]).split('\n') if c.strip()] if len(row)>7 else []
            self.hosts_data.append(h); self.table.insertRow(i); chk = QCheckBox(); chk.setChecked(True); w = QWidget(); l = QHBoxLayout(w); l.addWidget(chk); l.setAlignment(Qt.AlignCenter); l.setContentsMargins(0,0,0,0)
            self.table.setCellWidget(i, 0, w); self.table.setItem(i, 1, QTableWidgetItem(str(h.get('name','')))); self.table.setItem(i, 2, QTableWidgetItem(str(h.get('ip',''))))

    def toggle_all(self):
        if self.table.rowCount() == 0: return
        cur = self.table.cellWidget(0,0).layout().itemAt(0).widget().isChecked()
        for i in range(self.table.rowCount()): self.table.cellWidget(i,0).layout().itemAt(0).widget().setChecked(not cur)

    def reset_all_logs_and_tabs(self):
        self.global_console.clear()
        for widget in self.host_consoles.values():
            if isinstance(widget, QTextEdit): widget.clear()

    def run_process(self):
        selected = [self.hosts_data[i] for i in range(self.table.rowCount()) if self.table.cellWidget(i,0).layout().itemAt(0).widget().isChecked()]
        if not selected: return
        mode = self.combo.currentText()
        
        # Helper to set up common worker connections
        def start_worker(worker_obj):
            worker_obj.log_signal.connect(self.append_log)
            worker_obj.finished_signal.connect(self.on_worker_finished) # Data processing
            worker_obj.finished.connect(self.on_thread_finished) # Thread lifecycle & button reset
            self.active_workers.append(worker_obj)
            worker_obj.start()

        # === モード6: 自動診断 ===
        if "6:" in mode:
            if len(selected) != 1: return QMessageBox.warning(self, "エラー", "診断の出発点となる機器を1台だけ選択してください。")
            target_ip, ok = QInputDialog.getText(self, "自動診断設定", "宛先IPアドレスを入力:")
            if not ok or not target_ip: return
            
            self.btn_run.setEnabled(False); self.btn_cancel.setEnabled(True)
            worker = DiagnosticWorker(selected[0], target_ip, self.hosts_data)
            worker.path_data_signal.connect(self.visualize_path) # 経路図描画シグナル
            start_worker(worker); return

        # === モード7: 帯域モニター ===
        if "7:" in mode:
            if len(selected) != 1: return QMessageBox.warning(self, "エラー", "帯域モニターは1台のみ選択してください。")
            host = selected[0]
            try:
                # 簡易的なIF一覧取得
                dev = {'device_type': str(host.get('vendor','cisco')), 'host': host['ip'], 'username': host['user'], 'password': host['pw'], 'secret': host['en_pw']}
                with ConnectHandler(**dev) as net:
                    if ">" in net.find_prompt(): net.enable()
                    cmd = "show ip int brief"
                    if "juniper" in dev['device_type']: cmd = "show interfaces terse"
                    elif "huawei" in dev['device_type']: cmd = "display interface brief"
                    elif "hp" in dev['device_type'] or "aruba" in dev['device_type']: cmd = "show ip interface brief"
                    elif "arista" in dev['device_type']: cmd = "show ip interface brief"
                    elif "nec" in dev['device_type']: cmd = "show ip interface brief"
                    elif "allied" in dev['device_type']: cmd = "show ip interface brief"

                    res = net.send_command(cmd)
                ifaces = [l.split()[0] for l in res.splitlines() if l and not l.startswith(('Int', 'Name', ' ', 'PHY', 'Interface'))]
                iface, ok = QInputDialog.getItem(self, "IF選択", f"【{host['name']}】監視対象:", ifaces, 0, False)
                if not ok: return
            except Exception as e: return QMessageBox.critical(self, "エラー", f"接続失敗: {str(e)}")

            self.btn_run.setEnabled(False); self.btn_cancel.setEnabled(True); self.setup_ghost_tab(host['name'], iface)
            worker = TrafficGhostWorker(host, iface)
            worker.update_signal.connect(self.update_ghost_graph)
            start_worker(worker); return

        # === モード8: Crawler ===
        if "8:" in mode:
            if len(selected) != 1: return QMessageBox.warning(self, "エラー", "Crawlerの起点(Seed)となる機器を1台選択してください。")
            if not HAS_NETWORKX: return QMessageBox.critical(self, "エラー", "networkxライブラリがインストールされていません。")
            
            self.btn_run.setEnabled(False); self.btn_cancel.setEnabled(True)
            worker = CrawlerWorker(selected[0], self.hosts_data)
            worker.html_ready_signal.connect(self.open_topology_html)
            start_worker(worker); return

        # === モード9: Virtual Wiretap ===
        if "9:" in mode:
            if len(selected) != 1: return QMessageBox.warning(self, "エラー", "Wiretap対象の機器を1台選択してください。")
            host = selected[0]
            
            # --- インターフェース一覧を取得して選択させる ---
            try:
                dev = {'device_type': str(host.get('vendor','cisco')), 'host': host['ip'], 'username': host['user'], 'password': host['pw'], 'secret': host['en_pw']}
                with ConnectHandler(**dev) as net:
                    if ">" in net.find_prompt(): net.enable()
                    cmd = "show ip int brief"
                    if "juniper" in dev['device_type']: cmd = "show interfaces terse"
                    elif "huawei" in dev['device_type']: cmd = "display interface brief"
                    elif "linux" in dev['device_type']: cmd = "ip link show"
                    elif "hp" in dev['device_type'] or "aruba" in dev['device_type']: cmd = "show ip interface brief"
                    elif "arista" in dev['device_type']: cmd = "show ip interface brief"
                    
                    res = net.send_command(cmd)
                
                # パース
                ifaces = []
                for l in res.splitlines():
                    if "linux" in dev['device_type']:
                        m = re.match(r"\d+: ([^:@]+)", l)
                        if m: ifaces.append(m.group(1))
                    elif l and not l.startswith(('Int', 'Name', ' ', 'PHY', 'Interface')):
                        ifaces.append(l.split()[0])
                        
                iface, ok = QInputDialog.getItem(self, "Wiretap設定", f"【{host['name']}】キャプチャ対象:", ifaces, 0, False)
                if not ok: return
            except Exception as e: return QMessageBox.critical(self, "エラー", f"接続失敗: {str(e)}")
            # ------------------------------------------------------------------

            p_filter, ok = QInputDialog.getText(self, "Wiretap設定", "フィルタ (例: ip host 1.1.1.1):")
            if not ok: return
            duration, ok = QInputDialog.getInt(self, "Wiretap設定", "キャプチャ時間(秒):", 10, 5, 60)
            if not ok: return
            
            self.btn_run.setEnabled(False); self.btn_cancel.setEnabled(True)
            worker = WiretapWorker(host, iface, p_filter, duration)
            start_worker(worker); return

        # === 通常モード (0-5) ===
        self.btn_run.setEnabled(False); self.btn_cancel.setEnabled(True); self.btn_report.setEnabled(False)
        self.current_report_html = []
        self.active_workers = []
        self.full_mesh_matrix = {}

        if os.path.exists(SEARCH_FILE):
            with open(SEARCH_FILE, "r", encoding='utf-8-sig') as f: self.search_keywords = [l.strip() for l in f if l.strip()]
        else: self.search_keywords = []
        
        for host in selected:
            name = host['name']
            if name in self.host_consoles:
                con = self.host_consoles[name]
                if self.tabs.indexOf(con) == -1: self.tabs.addTab(con, name)
            else:
                con = ZoomableTextEdit(); con.setReadOnly(True)
                con.setStyleSheet("background-color:#1E1E1E; color:#FFFFFF; font-family:Consolas, monospace;")
                self.host_consoles[name] = con; self.tabs.addTab(con, name)

            worker = NetworkWorker(self.combo.currentText(), host, self.chk_show_log.isChecked(), self.chk_keyword_scan.isChecked(), self.search_keywords, selected, self.chk_compare_master.isChecked(), self.chk_save_master.isChecked(), self.teraterm_path)
            worker.log_signal.connect(self.append_log); worker.html_signal.connect(self.append_html); worker.finished_signal.connect(self.on_worker_finished); 
            worker.finished.connect(self.on_thread_finished) # Thread lifecycle
            worker.request_teraterm_path.connect(self.ask_teraterm_path)
            self.active_workers.append(worker); worker.start()

    def setup_ghost_tab(self, name, iface):
        t_title = f"MON: {name}"
        for i in range(self.tabs.count()):
            if self.tabs.tabText(i) == t_title: self.tabs.removeTab(i); break
        self.ghost_x, self.ghost_in, self.ghost_out = [], [], []
        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setStyleSheet("border: none; background-color: #1E1E1E;")
        self.canvas = MplCanvas(self, width=12, height=7, dpi=100)
        scroll.setWidget(self.canvas); self.tabs.addTab(scroll, t_title); self.tabs.setCurrentWidget(scroll)

    @Slot(str, float, float)
    def update_ghost_graph(self, time_str, in_m, out_m):
        self.ghost_x.append(time_str); self.ghost_in.append(in_m); self.ghost_out.append(out_m)
        if len(self.ghost_x) > 40: self.ghost_x.pop(0); self.ghost_in.pop(0); self.ghost_out.pop(0)
        self.canvas.axes.clear()
        self.canvas.axes.plot(self.ghost_x, self.ghost_in, label='Inbound (Mbps)', color='#00FF00', linewidth=2.5, alpha=0.8)
        self.canvas.axes.plot(self.ghost_x, self.ghost_out, label='Outbound (Mbps)', color='#5555FF', linewidth=2.5, alpha=0.8)
        self.canvas.axes.text(0.97, 0.95, f"NOW (Mbps)\nIN : {in_m:>7.2f}\nOUT: {out_m:>7.2f}", transform=self.canvas.axes.transAxes, color='white', fontsize=13, fontweight='bold', va='top', ha='right', family='Consolas', bbox=dict(facecolor='black', alpha=0.7, edgecolor='#555'))
        self.canvas.axes.set_facecolor('#181818'); self.canvas.axes.grid(True, color='#333', linestyle='--', lw=0.5)
        self.canvas.axes.set_xticks(range(len(self.ghost_x)))
        if len(self.ghost_x) > 0:
            step = max(1, len(self.ghost_x) // 8)
            self.canvas.axes.set_xticklabels([self.ghost_x[i] if i % step == 0 or i == len(self.ghost_x)-1 else "" for i in range(len(self.ghost_x))], color='#AAA', fontsize=9, rotation=45)
        self.canvas.axes.tick_params(axis='y', colors='#AAA', labelsize=9); self.canvas.axes.legend(loc='upper left', facecolor='#222', labelcolor='white')
        self.canvas.axes.set_ylim(bottom=0); self.canvas.fig.tight_layout(); self.canvas.draw()

    @Slot(list)
    def visualize_path(self, path_data):
        if not HAS_NETWORKX: return
        t_title = "Path Visualizer"
        for i in range(self.tabs.count()):
            if self.tabs.tabText(i) == t_title: self.tabs.removeTab(i); break
        
        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setStyleSheet("border: none; background-color: #1E1E1E;")
        canvas = MplCanvas(self, width=12, height=7, dpi=100)
        scroll.setWidget(canvas); self.tabs.addTab(scroll, t_title); self.tabs.setCurrentWidget(scroll)

        G = nx.DiGraph()
        labels = {}; node_colors = []; edge_colors = []
        
        for i, p in enumerate(path_data):
            node_name = p['node']
            G.add_node(node_name)
            
            lbl = node_name
            if p.get('reason'):
                lbl += f"\n[{p['reason']}]"
            labels[node_name] = lbl
            
            color = "#00FF00" if p.get('status') == "OK" else "#FF0000"
            node_colors.append(color)
            
            if i > 0:
                prev = path_data[i-1]['node']
                curr = node_name
                G.add_edge(prev, curr)
                edge_colors.append(color)

        pos = nx.spring_layout(G, seed=42) if len(path_data) > 2 else nx.planar_layout(G)
        
        nx.draw_networkx_nodes(G, pos, ax=canvas.axes, node_color=node_colors, node_size=2000)
        nx.draw_networkx_edges(G, pos, ax=canvas.axes, edge_color=edge_colors, width=2, arrows=True)
        nx.draw_networkx_labels(G, pos, ax=canvas.axes, labels=labels, font_color='white', font_weight='bold', font_size=9)
        canvas.axes.set_axis_off(); canvas.draw()

    @Slot(str)
    def open_topology_html(self, path):
        reply = QMessageBox.question(self, "トポロジー生成完了", 
                                   f"インタラクティブなトポロジーマップを生成しました。\nブラウザで開きますか？\n\n保存先: {path}",
                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        if reply == QMessageBox.Yes:
            QDesktopServices.openUrl(QUrl.fromLocalFile(path))

    def stop_workers(self):
        for w in self.active_workers:
            if isinstance(w, TrafficGhostWorker): w.stop()
            
            # [Fix] 安全な停止手順
            if isinstance(w, NetworkWorker):
                w.stop() # フラグを立ててプロセスを停止
                
                # スレッドが自然に終わるのを最大2秒待つ
                if not w.wait(2000):
                    # タイムアウトしたら強制終了（最終手段）
                    w.terminate()
            elif w.isRunning():
                w.terminate()
                w.wait()
        
        self.active_workers = []; self.btn_run.setEnabled(True); self.btn_cancel.setEnabled(False); self.append_log("GLOBAL", "\n[!!!] キャンセルされました。", "#FF5555")
        QMessageBox.information(self, "通知", "実行中の処理を中止しました")

    @Slot()
    def ask_teraterm_path(self):
        if self.teraterm_path: return
        p, _ = QFileDialog.getOpenFileName(self, "TeraTermマクロ(ttpmacro.exe)を選択", "C:\\", "Executable (*.exe)")
        if p:
            self.teraterm_path = p
            self.append_log("GLOBAL", f"TeraTermパスを設定しました: {p}", "#00FF00")
            QMessageBox.information(self, "設定完了", "パスを設定しました。再度実行してください。")

    @Slot(str, list, dict)
    def on_worker_finished(self, name, report, mesh):
        if report: self.current_report_html.extend(report)
        if mesh: self.full_mesh_matrix[name] = mesh
        # NOTE: cleanup moved to on_thread_finished to avoid race conditions

    @Slot()
    def on_thread_finished(self):
        # Filter out threads that are actually finished
        self.active_workers = [w for w in self.active_workers if w.isRunning()]
        
        if not self.active_workers:
            self.btn_run.setEnabled(True)
            self.btn_cancel.setEnabled(False)
            
            if "5:" in self.combo.currentText(): 
                self.generate_mesh_report()
            
            self.append_log("GLOBAL", "\n--- 全ての処理が完了しました ---", "#00FF00")
            self.btn_report.setEnabled(True if self.current_report_html else False)

    def generate_mesh_report(self):
        hosts = sorted(self.full_mesh_matrix.keys()); t_name = "疎通マトリックス"
        for i in range(self.tabs.count()):
            if self.tabs.tabText(i) == t_name: self.tabs.removeTab(i); break
        html = f'<div style="text-align: left; margin-left: 0; margin-top: 20px;">'
        html += f'<div style="color:#FFFF00; font-weight:bold; margin-bottom: 10px; font-family:sans-serif;">疎通マトリックス結果</div>'
        html += '<table border="1" style="border-collapse:collapse; margin-left:0; color:#eee; background:#222; text-align:center; width: auto; font-family:Consolas, monospace;">'
        html += '<tr style="background:#444;"><th>FROM \\ TO</th>' + "".join([f'<th style="padding:5px 10px;">{h}</th>' for h in hosts]) + '</tr>'
        for src in hosts:
            row = f'<tr><td style="background:#444; padding:5px 10px;"><b>{src}</b></td>'
            for dst in hosts:
                st = self.full_mesh_matrix[src].get(dst, "-"); color = "#00FF00" if st == "OK" else "#FF5555" if st == "NG" else "#888"
                row += f'<td style="color:{color}; font-weight:bold; padding:5px 10px;">{st}</td>'
            html += row + '</tr>'
        html += '</table></div>'
        con = ZoomableTextEdit(); con.setReadOnly(True); con.setStyleSheet("background-color:#1E1E1E; color:#FFFFFF; font-family:Consolas, monospace;"); con.append(html); self.tabs.addTab(con, t_name); self.tabs.setCurrentWidget(con); self.current_report_html.append(html)

    def save_report(self):
        f_p, _ = QFileDialog.getSaveFileName(self, "レポート保存", os.path.join(REPORT_DIR, f"Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"), "HTML Files (*.html)")
        if f_p:
            with open(f_p, "w", encoding='utf-8') as f: f.write(f'<html><body style="background:#111; color:#eee; padding:30px; font-family:sans-serif;">{"".join(self.current_report_html)}</body></html>')

    @Slot(str, str, str)
    def append_log(self, name, text, color):
        l = f'<span style="color:{color}; white-space:pre-wrap;">{text}</span>'; self.global_console.append(l)
        if name in self.host_consoles: self.host_consoles[name].append(l)

    @Slot(str, str)
    def append_html(self, name, html):
        self.global_console.append(html)
        if name in self.host_consoles: self.host_consoles[name].append(html)

if __name__ == "__main__":
    app = QApplication(sys.argv); app.setStyle("Fusion"); window = NetVerifyGUI(); window.show(); sys.exit(app.exec())
